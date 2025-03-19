import openpyxl
import json
import os

def get_excel_data(excel_file_path, sheet_name=None):
    """
    Reads an Excel file and extracts cell formatting, sizes, etc. 
    Returns a dictionary ready to be saved as JSON.
    """
    wb = openpyxl.load_workbook(excel_file_path, data_only=False)
    
    # Use the first sheet if sheet_name not provided
    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[sheet_name]

    # Gather row heights
    row_heights = {}
    for row_idx in range(ws.min_row, ws.max_row + 1):
        if ws.row_dimensions[row_idx].height is not None:
            row_heights[row_idx] = ws.row_dimensions[row_idx].height
        else:
            # Default row height if none is specified
            row_heights[row_idx] = 12.75  # typical default in Excel

    # Gather column widths
    col_widths = {}
    for col_idx in range(ws.min_column, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if ws.column_dimensions[col_letter].width is not None:
            col_widths[col_idx] = ws.column_dimensions[col_letter].width
        else:
            # Default column width if none is specified
            col_widths[col_idx] = 8.43  # typical default in Excel

    # Gather cell data
    cells_data = []
    for row_idx in range(ws.min_row, ws.max_row + 1):
        for col_idx in range(ws.min_column, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # -------------------------
            # Font color (only if type == 'rgb')
            # -------------------------
            font_color = None
            if cell.font and cell.font.color:
                if cell.font.color.type == "rgb":
                    font_color = cell.font.color.rgb  # e.g. 'FFFF0000'
            
            # -------------------------
            # Fill color (only if type == 'rgb')
            # -------------------------
            fill_color = None
            if cell.fill and cell.fill.fgColor:
                if cell.fill.fgColor.type == "rgb":
                    fill_color = cell.fill.fgColor.rgb

            # Cell Font data
            font_name = cell.font.name if cell.font else "Arial"
            font_size = cell.font.sz if cell.font else 11
            bold = cell.font.b if cell.font else False
            italic = cell.font.i if cell.font else False
            underline = cell.font.u if cell.font else None

            # Cell border styles (convert to string so JSON can handle them)
            borders = {
                'left':   str(cell.border.left.style),
                'right':  str(cell.border.right.style),
                'top':    str(cell.border.top.style),
                'bottom': str(cell.border.bottom.style),
            }

            # Build cell dict
            cell_dict = {
                'row': row_idx,
                'col': col_idx,
                'value': str(cell.value) if cell.value is not None else '',
                'font': {
                    'name': font_name,
                    'size': font_size,
                    'bold': bold,
                    'italic': italic,
                    'underline': underline,
                    'color': font_color  # either "FFAABBCC" or None
                },
                'fill': {
                    'color': fill_color  # either "FFAABBCC" or None
                },
                'borders': borders
            }
            cells_data.append(cell_dict)

    # Combine all into a single data structure
    data = {
        'sheet_name': ws.title,
        'row_heights': row_heights,
        'column_widths': col_widths,
        'cells': cells_data,
    }

    return data

def save_json(data, json_file_path):
    """
    Saves the extracted Excel data into a JSON file.
    """
    with open(json_file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)

def main():
    excel_path = r"C:\tmp\LegendTransferTemplate.xlsx"
    json_path  = r"C:\tmp\save_data.json"

    excel_data = get_excel_data(excel_path, sheet_name="Sheet1")
    save_json(excel_data, json_path)
    print(f"Data successfully saved to {json_path}")

if __name__ == "__main__":
    main()
